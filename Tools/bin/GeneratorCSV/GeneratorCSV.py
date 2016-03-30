#-*- coding: utf-8 -*-
#coding:utf-8
import sys , getopt
import os , re , string
import time , datetime
import binascii 
import collections  
import xml.dom.minidom
import time,datetime
import socket
import csv

#g_platform = "LINUX"
g_platform = "WIN64"
if g_platform == "WIN64":
	import ctypes
	
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook

from xml.dom import minidom , Node 

from xml.etree.ElementTree import ElementTree
from xml.etree.ElementTree import Element
from xml.etree.ElementTree import SubElement as SE
#sys.reload()
#sys.setdefaultencoding("utf-8") 
#sys.setdefaultencoding("cp936") 

g_yellow = ""
g_yellow_h = ""
g_green = ""
g_green_h = ""
g_red = ""
g_original = ""
g_cyan = ""
g_stdInputHandle = -10
g_stdOutputHandle = -11
g_stdErrorHandle = -12

#windows下的颜色.
FOREGROUND_BLACK = 0x0
FOREGROUND_BLUE = 0x01
FOREGROUND_GREEN = 0x02
FOREGROUND_RED = 0x04
FOREGROUND_INTENSTITY = 0x08

BACKGROUND_BLUE = 0x20
BACKGROUND_GREEN = 0x30
BACKGROUND_RED = 0x40
BACKGROUND_INTENSTITY = 0x80
			
g_xlsImportPath = ""
g_xlsExportCSVPath = ""
g_xlsExportCPPPath = ""

g_conditionConfigName = "_ConditionConfig"
g_commonDataName = "_CommonData"
g_globalFuncName = "Global"
g_xlsDeleteRecord = []
g_xlsRecords = collections.OrderedDict()
g_xlsConditionRecords = collections.OrderedDict()
	
g_commonDatas = collections.OrderedDict()  # 这里用来记录所有的共有变量的值

g_serverConditions = collections.OrderedDict()  # 这里用来记录所有的条件
g_serverActions = collections.OrderedDict()
g_clientConditions = collections.OrderedDict()
g_clientActions = collections.OrderedDict()
g_serverConditionCol = 1
g_serverActionCol = 2			
g_clientConditionCol = 3
g_clientActionCol = 4
g_serverExpression = collections.OrderedDict()
g_clientExpression = collections.OrderedDict()
g_conditionObjName = "pConfig"		# 这个是当处理条件的时候.如果条件是本表内的,使用的统一名字.
g_actionNameSplit = "("				# 执行时的名字.用这个来区分
g_expressionPrefix = "Run"
g_and = "&&"	#没有或和非运算.如果想用这个.多写几个条件.
g_great = ">"
g_greatEqual = ">="
g_less = "<"
g_lessEqual = "<="
g_equal = "=="
g_notEqual = "!="


g_configPrefix = "S"
g_loadConfigSuffix = "Load"
g_xlsNamespace = "Config"
g_conditionTag = "condition"
g_actionTag = "action"

g_int32Type = "INT32"
g_int64Type = "INT64"
g_boolType = "bool"
g_doubleType = "double"
g_stringType = "std::string"
g_dateType = "Timer::Date"
g_int32ArrayType = "std::vector<INT32>" 
g_int64ArrayType = "std::vector<INT64>"
g_boolArrayType = "std::vector<bool>"
g_doubleArrayType = "std::vector<double>"
g_stringArrayType = "std::vector<std::string>"
g_dateArrayType = "std::vector<Timer::Date>"
g_conditionType = "Condition"
g_structType = 1
g_structArrayType = 2
	
g_int32Func = "GetInt32"
g_int64Func = "GetInt64"
g_boolFunc = "GetBool"
g_doubleFunc = "GetDouble"
g_stringFunc = "GetString"
g_dateFunc = "GetDate"
g_int32ArrayFunc = 6
g_int64ArrayFunc = 7
g_boolArrayFunc = 8
g_doubleArrayFunc = 9
g_stringArrayFunc = 10
g_structFunc = 11
g_structArrayFunc = 12
g_dateArrayFunc = 13

oneTab = "\t"
twoTab = oneTab + "\t"
threeTab = twoTab + "\t"
fourTab = threeTab + "\t"
fiveTab = fourTab + "\t"
sixTab = fiveTab + "\t"
sevenTab = sixTab + "\t"
eightTab = sevenTab + "\t"

g_rowComent = 0 # 注释行
g_rowType = 1	# 类型行
g_rowName = 2 	# 为类型名字行
g_rowCS = 3   	# 第三行为是否是服务器还是客户端代码
g_cellID = 0  	# 第一列为ID列
g_cellName = 0 	# 条件列的   条件名字
g_cellCS = 1 	# 条件列的   服务器还是客户端代码.
g_rowDataStart=4# 从这行开始就是数据了.

class Expression(object):
	"""docstring for Expression"""
	def __init__(self, arg):
		super(Expression, self).__init__()
		self.condition = None
		self.actions = []
		
class Condition(object):
	"""docstring for Condition"""
	def __init__(self, arg):
		super(Condition, self).__init__()
		self.leftvalue = ""
		self.symbol = ""
		self.rightvalue = ""
		
class Action(object):
	"""docstring for Action"""
	def __init__(self, arg):
		super(Action, self).__init__()
		self.name = ""
		self.args = []
		

def start(): 
	LogOutInfo("start generate csv.\n")   
	CreateExportPathFiles()
	DeleteExportPathFiles()
	GenerateCSVFromXLS()
	LogOutInfo("generate CSV finished.\n") 
			
	LogOutInfo("start generate CPP.\n")   	
	GenerateCPP()
	LogOutInfo("generate CPP finished.\n") 

def GenerateCSVFromXLS():
	root = g_xlsImportPath
	
	files = Search(root ,'.xlsx')
	for result in files:
		LogOutInfo("filename:" , result);
		Xlsx2CSV(result)

	CheckRecords()
	GenerateCSV()		
	HandleConditionConfig()
	HandleSheetCondition()

def Xlsx2CSV(filepath):
	dirout = g_xlsExportCSVPath
	dirout = dirout + os.sep  #新路径名称
	try:
		# 一个表中的所有sheet输出到一个csv文件中,所以要保证格式一致.文件名用xlsx文件名
		filename = os.path.basename(filepath) #获取文件名
		filename = os.path.splitext(filename.replace(' ', '_'))[0]
		if filename.find('#') >= 0:
			filename = filename.replace('#' , '')
			LogOutInfo("delete filename.filename=" , filename )
			g_xlsDeleteRecord.append(filename)

		#csv_filename = '{xlsx}.tabcsv'.format(xlsx=filename)		
		#dirfileout  = dirout + csv_filename
		#LogOutDebug("dirfileout" , dirfileout )
		#csv_file = open(dirfileout , 'w' , newline='')
		
		#QUOTE_MINIMAL QUOTE_NONE所有的都要加引号.
		#csv_file_writer = csv.writer(csv_file , delimiter='	', quotechar='"', quoting=csv.QUOTE_ALL)
		
		id_list = []			# 用于重复的ID去除
		delete_col_list = []	# 不读取需要删除的列
		cur_sheet_index = 0
		xlsx_file_reader = load_workbook(filepath)
		g_xlsRecords[filename] = collections.OrderedDict()
		conditionContainer = collections.OrderedDict() #这个专门处理条件表达式的.
		for sheet in xlsx_file_reader.get_sheet_names():			
			cur_rows_index = 0
			sheet_ranges = xlsx_file_reader[sheet]
			for row in sheet_ranges.rows:
				if cur_sheet_index >= 1 and cur_rows_index < g_rowDataStart + 1: #当第二个sheet的时候前4行是不读取的.
					cur_rows_index = cur_rows_index + 1
					continue
				
				row_container = [] 
				cur_cell_index = 0
				for cell in row:		
					Str = ""	
					if type(cell.value) == type(None):
						if cur_cell_index == g_cellID:
							break		
						Str = ""	
#						LogOutError("error parase filepath" , filepath , "  cur_sheet " , sheet , "  cur_rows_index " , cur_rows_index ,"  cur_cell_index " , cur_cell_index , "  type(cell.value) " , type(cell.value))
					else:
						Str = cell.value
						if type(cell.value) != str:
							Str = str(cell.value)
						else:
							Str = Str.encode('gbk').decode('gbk')

					if cur_rows_index == g_rowComent and len(Str) > 0 and Str.lower()[0] == '#':  # 跳过这一列
						#LogOutDebug("delete_col_list" , delete_col_list)
						delete_col_list.append(cur_cell_index)
					if cur_cell_index in delete_col_list:
						cur_cell_index = cur_cell_index + 1
						#LogOutDebug("cur_cell_index continue" , cur_cell_index)
						continue

					if cur_cell_index == g_cellID and len(Str) > 0 and Str.lower()[0] == '#':  # 跳过这一行
						#LogOutDebug("cur_cell_index break " , cur_cell_index)
						cur_cell_index = cur_cell_index + 1
						break

					if len(Str) >= 0:			#有些策划没填的数据为空.也需要记录.
						if cur_rows_index < 4:
							Str = ''.join([x for x in Str if x != " "]) 
						if cur_cell_index == g_cellID:		# 第一列是ID.需要判定是否有重复的ID
							if Str in id_list:
								LogOutDebug("repeat id \'" , Str , "\' in \'" , filepath , " \' file")
								#LogOutError("repeat id \'" , Str , "\' in \'" , filepath , " \' file")
							id_list.append(Str)
						if cur_rows_index == g_rowType and Str.lower() == g_conditionType.lower():
							conditionContainer[cur_cell_index] = []  #这个用来记录这一列的条件表达式.
						else:
							if cur_cell_index in conditionContainer:
								if len(Str) > 0:  # 记录条件的时候因为不是一一对应的,所以当为空的时候不需要记录.
									conditionContainer[cur_cell_index].append(Str)
									RemovListNewLine(conditionContainer[cur_cell_index])
									#LogOutDebug("conditionContainer:" , Str)
							else:
								row_container.append(Str)

					cur_cell_index = cur_cell_index + 1

				if len(row_container) >= 1:	
					RemovListNewLine(row_container)
					g_xlsRecords[filename][cur_rows_index] = row_container
					#csv_file_writer.writerow(row_container)
					cur_rows_index = cur_rows_index + 1   #这里是为了去除一些空行.
					#LogOutDebug("cell.row_container:" , row_container)
			cur_sheet_index = cur_sheet_index + 1	
		g_xlsConditionRecords[filename] = conditionContainer
		#LogOutDebug("g_xlsConditionRecords:" , g_xlsConditionRecords[filename])
		#csv_file.close()
	except Exception as e:
		LogOutError(e)
		
def CheckRecords():
	for sheet , item in g_xlsRecords.items(): 	#读取sheet
		#LogOutDebug("sheet" , sheet)
		for row , rowItem in item.items():		#读取每一行
			if row == g_rowName:				#命名重名检测,
				sameName = []
				for col , colItem in enumerate(rowItem):	#读取每一列
					colItemName = colItem
					item_type = GetType(g_xlsRecords[sheet][g_rowType][col])
										
					if colItem[0] == '[' or colItem[0].isdigit():
						LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , colItem , "use error name .")
					if item_type == g_structArrayType:
						if colItem.find('[') <= 0 or colItem.find(']') <= 0:
							LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , colItem , "error  .no \'[\' or \']\'.")
					
					if item_type == g_structArrayType or item_type == g_structType:
						npos = colItem.find('[')
						colItemName = colItem[0 : npos]
					
					if colItemName in sameName:
						LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , colItemName , " use one same name .")
					else:
						sameName.append(colItemName)						
			elif row == g_rowType:		#类型
				for col , colItem in enumerate(rowItem):	#读取每一列
					item_type = GetType(colItem)					
				
			elif row == g_rowCS:	
				pass
			elif row == g_rowComent:	
				pass	
			else:
				for col , colItem in enumerate(rowItem):	#读取每一列
					item_type = GetType(g_xlsRecords[sheet][g_rowType][col])	
					CheckDataType(item_type , sheet , row , col , colItem)
					#LogOutDebug("CheckDataType,row=" + str(row) + ":item_type=" , item_type)
				#LogOutDebug("g_xlsRecords,[row = " + str(row) + ":" , g_xlsRecords[sheet][row])
					
	for sheet , item in g_xlsConditionRecords.items():
		for col , colItem in item.items():		# 读取每一列
			sameName = []						# 命名重名检测,
			for row , rowItem in enumerate(colItem):		# 读取每一行
				if row == 0:	# 第一行是名字
					sameName.append(rowItem)
				elif rowItem in sameName:
					LogOutError("sheet=" , sheet  , " :row=" , row , " :col=" , col , " item=" , rowItem , " use one same name .")

			for row , rowItem in enumerate(colItem):	#读取每一列
				CheckDataType(g_conditionType , sheet , row , col , rowItem)
				#LogOutDebug("g_xlsRecords[sheet][row][col]:" , g_xlsRecords[sheet][row][col])
				
def HandleConditionConfig():
	rowItems = g_xlsRecords[g_conditionConfigName]
	if len(rowItems) != 0:
		for row , rowItem in rowItems.items():				# 读取每一行
			for col , colItem in enumerate(rowItem):		# 读取每一列
				if row >= g_rowDataStart: 					# 从这行开始是数据	
					#LogOutInfo("row:" , row , " col:" , col , " colItem:" , colItem)	
					if len(colItem) == 0:
						continue			
					if col == g_serverConditionCol:
						objects = colItem.split('.')		# 这里至少是2个.
						last = objects[len(objects) - 1].upper()
						if last not in g_serverConditions:
							g_serverConditions[last] = []
						for index , obj in enumerate(objects):
							if index != len(objects) - 1:
								g_serverConditions[last].append(obj)
					if col == g_serverActionCol:
						objects = colItem.split('.')		# 这里至少是2个.
						last = objects[len(objects) - 1].upper()
						if last not in g_serverActions:
							g_serverActions[last] = []
						for index , obj in enumerate(objects):
							if index != len(objects) - 1:
								g_serverActions[last].append(obj)
					if col == g_clientConditionCol:
						objects = colItem.split('.')		# 这里至少是2个.
						last = objects[len(objects) - 1].upper()
						if last not in g_clientConditions:
							g_clientConditions[last] = []
						for index , obj in enumerate(objects):
							if index != len(objects) - 1:
								g_clientConditions[last].append(obj)
					if col == g_clientActionCol:
						objects = colItem.split('.')		# 这里至少是2个.
						last = objects[len(objects) - 1].upper()
						if last not in g_clientActions:
							g_clientActions[last] = []
						for index , obj in enumerate(objects):
							if index != len(objects) - 1:
								g_clientActions[last].append(obj)
		#LogOutDebug("g_serverConditions:" , g_serverConditions)			
		#LogOutDebug("g_serverActions:" , g_serverActions)
		#LogOutDebug("g_clientConditions:" , g_serverConditions)
		#LogOutDebug("g_clientActions:" , g_clientActions)
	else:
		LogOutError("HandleConditionConfig error." , g_conditionConfigName , " no data")

def HandleSheetCondition():	
	for sheet , item in g_xlsConditionRecords.items():
		g_serverExpression[sheet] = collections.OrderedDict()			# 为每一个sheet创建一系列的条件表达式
		for col , colItem in item.items():		# 读取每一列
			bConditionStart = False
			expression = None
			cellName = ""
			for row , rowItem in enumerate(colItem):		# 读取每一行
				item = RemoveSpecialWord(rowItem)
				if row == g_cellName:
					cellName = g_expressionPrefix + rowItem
					#LogOutDebug("cellName:" , cellName , "g_serverExpression[sheet]:" , g_serverExpression[sheet])
					g_serverExpression[sheet][cellName] = []	# 这里用来存每一个表达式
				if row != g_cellName and row != g_cellCS:
					tagitem = item.lower()[0:item.lower().find(':')]
					itemContent = item.lower()[item.lower().find(':') + 1:]
					#LogOutDebug("condition item:" , item , " tagitem=" , tagitem , " itemContent " , itemContent)
					if tagitem.lower() == g_conditionTag.lower():
						if not bConditionStart:
							if expression != None:
								g_serverExpression[sheet][cellName].append(expression)
							expression = HandleConditionExpression(True , sheet , itemContent)
							expression.actions = []
							bConditionStart = True
					elif tagitem.lower() == g_actionTag.lower():
						if bConditionStart:
							action = HandleAction(True , itemContent)
							expression.actions.append(action)							
					else:
						LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , item , " condition tag error.")							
			
			if expression != None and cellName != "":
				g_serverExpression[sheet][cellName].append(expression)	
		LogOutDebug("g_serverExpression[sheet]:" , g_serverExpression[sheet])	
def HandleConditionExpression(bServer , sheet , itemContent):	
	expression = Expression
	conditions = itemContent.split("&&")
	expression.condition = Condition
	for index , item in enumerate(conditions):
		if item.find(g_great) >= 0:
			expression.condition.leftvalue = CheckConditionValue(bServer , sheet , item.split(g_great)[0])
			expression.condition.symbol = g_great
			expression.condition.rightvalue = CheckConditionValue(bServer , sheet , item.split(g_great)[1])
		if item.find(g_greatEqual) >= 0:
			expression.condition.leftvalue = CheckConditionValue(bServer , sheet , item.split(g_greatEqual)[0])
			expression.condition.symbol = g_greatEqual
			expression.condition.rightvalue = CheckConditionValue(bServer , sheet , item.split(g_greatEqual)[1])
		if item.find(g_less) >= 0:
			expression.condition.leftvalue = CheckConditionValue(bServer , sheet , item.split(g_less)[0])
			expression.condition.symbol = g_less
			expression.condition.rightvalue = CheckConditionValue(bServer , sheet , item.split(g_less)[1])
		if item.find(g_lessEqual) >= 0:
			expression.condition.leftvalue = CheckConditionValue(bServer , sheet , item.split(g_lessEqual)[0])
			expression.condition.symbol = g_lessEqual
			expression.condition.rightvalue = CheckConditionValue(bServer , sheet , item.split(g_lessEqual)[1])
		if item.find(g_equal) >= 0:
			expression.condition.leftvalue = CheckConditionValue(bServer , sheet , item.split(g_equal)[0])
			expression.condition.symbol = g_equal
			expression.condition.rightvalue = CheckConditionValue(bServer , sheet , item.split(g_equal)[1])
		if item.find(g_notEqual) >= 0:
			expression.condition.leftvalue = CheckConditionValue(bServer , sheet , item.split(g_notEqual)[0])
			expression.condition.symbol = g_notEqual
			expression.condition.rightvalue = CheckConditionValue(bServer , sheet , item.split(g_notEqual)[1])

	return expression

def CheckConditionValue(bServer , sheet ,  value):
	if bServer:
		if value.upper() in g_serverConditions:
			result = ""
			for index , item in enumerate(g_serverConditions[value.upper()]):
				result += item
				result += "::"
			result += value.upper()
			return result
		if value in g_xlsRecords[sheet][g_rowName]:
			result = g_conditionObjName + "->" + value
			return result
		if value in g_xlsRecords[sheet][g_rowName]:  # todo如果在commondata中该如何处理.
			result = g_conditionObjName + "->" + value
			return result
	else:
		if value.upper() in g_clientConditions:
			result = ""
			for index , item in enumerate(g_clientConditions[value.upper()]):
				result += item
				result += "::"
			result += value.upper()
			return result
		if value in g_xlsRecords[sheet][g_rowName]:
			result = g_conditionObjName + "->" + value
			return result
		if value in g_xlsRecords[sheet][g_rowName]:  # todo如果在commondata中该如何处理.
			result = g_conditionObjName + "->" + value
			return result
	
def HandleAction(bServer , itemContent):	
	action = Action
	actions = itemContent.split(g_actionNameSplit)
	if len(actions) > 0:
		action.name = CheckActionValue(bServer , actions[0])
		RemoveSpecialWord(actions[1])
		action.args = actions[1].split(',')
			 
	return action

def CheckActionValue(bServer ,  value):
	if bServer:
		if value.upper() in g_serverActions:
			result = ""
			for index , item in enumerate(g_serverActions[value.upper()]):
				result += item
				result += "::"
			result += value.upper()
			return result
		else:
			LogOutError("CheckActionValue error." , value , " g_serverActions:" , g_serverActions)
	else:
		if value.upper() in g_clientActions:
			result = ""
			for index , item in enumerate(g_clientActions[value.upper()]):
				result += item
				result += "::"
			result += value.upper()
			return result
		else:
			LogOutError("CheckActionValue error." , value)

def GenerateCSV():
	
	dirout = g_xlsExportCSVPath
	dirout = dirout + os.sep  #新路径名称
	try:
		for sheet , item in g_xlsRecords.items(): 	#读取sheet
			filename = sheet
			csv_filename = '{xlsx}.tabcsv'.format(xlsx=filename)		
			dirfileout  = dirout + csv_filename
			#LogOutDebug("dirfileout" , dirfileout )
			csv_file = open(dirfileout , 'w' , newline='')
			
			#QUOTE_MINIMAL QUOTE_NONE所有的都要加引号.
			csv_file_writer = csv.writer(csv_file , delimiter='	', quotechar='"', quoting=csv.QUOTE_ALL)

			for row , rowItem in item.items():	#读取每一行
				csv_file_writer.writerow(rowItem)

			csv_file.close()
	except Exception as e:
		LogOutError(e)

def GenerateCPP(): 
	GenerateConfigManagerHeader()
	for sheet , item in g_xlsRecords.items():
		GenerateConfigLoadHeader(sheet , item[g_rowType] , item[g_rowName] , item[g_rowComent])
		GenerateConfigLoadCpp(sheet , item[g_rowType] , item[g_rowName] , item[g_rowComent])
		GenerateConfigHeader(sheet , item[g_rowType] , item[g_rowName] , item[g_rowComent])
		GenerateConfigCpp(sheet , item[g_rowType] , item[g_rowName] , item[g_rowComent])
	GenerateConfigManagerCPP()

def GenerateStructData(fileWrite , types , datas , comments):	
	for index , item in enumerate(types):
		item_type = GetType(item)
		if item_type == g_structType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			#LogOutDebug("name=" , structName , ":structDatas " , structDatas , " size=" , len(structDatas))
			fileWrite.write("\n" + twoTab + "//" + comments[index] + "\n") 
			fileWrite.write(twoTab + "struct " + g_configPrefix + structName + "\n");
			fileWrite.write(twoTab + "{\n");
			childItems = item.split(',')
			if len(childItems) != len(structDatas):
				LogOutError("parase struct error.maybe no detail name. structName=" , structName , ":childItems " , childItems , "size=" , len(childItems) , ":structDatas " , structDatas , " size=" , len(structDatas))
			for indexChild , childItem in enumerate(childItems):
				#LogOutDebug("name=" , structName , ":structDatas=" , structDatas , ":childItem=" , childItem)
				fileWrite.write(threeTab + GetType(childItem) + GetTypeTab(childItem) + structDatas[indexChild] + ";\n")
			fileWrite.write(twoTab + "}" + structName + ";\n");
		elif item_type == g_structArrayType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			fileWrite.write("\n" + twoTab + "//" + comments[index] + "\n") 
			fileWrite.write(twoTab + "struct " + g_configPrefix + structName + "\n");
			fileWrite.write(twoTab + "{\n");
			item = item.replace('[' , '').replace(']' , '')
			childItems = item.split(',')
			if len(childItems) != len(structDatas):
				LogOutError("parase struct error.invalid size . name=" , structName , ":childItems " , childItems , "size=" , len(childItems) , ":structDatas " , structDatas , " size=" , len(structDatas))
			for indexChild , childItem in enumerate(childItems):
				fileWrite.write(threeTab + GetType(childItem) + GetTypeTab(childItem) + structDatas[indexChild] + ";\n")
			fileWrite.write(twoTab + "}" + ";\n");
			fileWrite.write(twoTab + "std::vector<" + g_configPrefix + structName + ">" + twoTab + "vec" + structName + ";\n");
		else:
			fileWrite.write(twoTab + item_type + GetTypeTab(item) + oneTab + datas[index] + ";"  + oneTab)
			fileWrite.write("//" + comments[index] + "\n") 

def GenerateConfigLoadHeader(filename , types , datas , comments):
	filename = filename + g_loadConfigSuffix
	outputPath = g_xlsExportCPPPath  + os.sep + filename + ".h"
	if os.path.exists(outputPath): 
		os.remove(outputPath)

	fileWrite = open(outputPath , "a")
	
	#以下是生成数据定义
	WriteFileDescription(fileWrite , filename + ".h" , "csv配置文件")
	fileWrite.write("#ifndef __" + filename + "_define_h__\n")
	fileWrite.write("#define __" + filename + "_define_h__\n") 
	fileWrite.write("#include \"CUtil/inc/Common.h \"\n\n") 
	fileWrite.write("#include \"Timer/inc/Date.h \"\n\n") 
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n") 
	fileWrite.write(oneTab + "struct " + g_configPrefix + filename + "\n") 
	fileWrite.write(oneTab + "{\n") 
	GenerateStructData(fileWrite , types , datas , comments)
	fileWrite.write(oneTab + "};\n\n\n") 
			
	#以下是生成导入数据接口
	fileWrite.write(oneTab + "class " + filename + "\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "typedef std::vector<" + g_configPrefix + filename + "> CollectionConfigsT;\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "bool LoadFrom(const std::string& filename);\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + g_configPrefix + filename + " & Get(size_t row);\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "inline size_t Count()")
	fileWrite.write("{ ")
	fileWrite.write("return m_vtConfigs.size(); ")
	fileWrite.write("}\n\n")
	fileWrite.write(oneTab + "private:\n")
	fileWrite.write(twoTab + "CollectionConfigsT m_vtConfigs;\n")	
	fileWrite.write(oneTab + "};\n") 


	fileWrite.write("}\n\n" + "#endif// end" + "  __" + filename + "_define_h__\n") 

	fileWrite.close()	  
	
def GenerateConfigLoadCpp(filename , types , datas , comments):
	filename = filename + g_loadConfigSuffix
	outputPath = g_xlsExportCPPPath  + os.sep + filename + ".cpp"
	if os.path.exists(outputPath): 
		os.remove(outputPath)

	fileWrite = open(outputPath , "a")
	
	WriteFileDescription(fileWrite , filename + ".cpp" , "csv读取文件实现")
	fileWrite.write("#include \"" + filename + ".h\"\n") 
	fileWrite.write("#include \"CUtil/inc/CUtil.h\"\n") 
	fileWrite.write("#include \"CUtil/inc/CSVReader.h\"\n\n") 
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n") 
	fileWrite.write(oneTab + "bool " + filename + "::LoadFrom(const std::string & filepath)\n") 
	fileWrite.write(oneTab + "{\n") 
	
	fileWrite.write(twoTab + "CUtil::CSVReader csv;\n") 
	fileWrite.write(twoTab + "if(csv.Load(filepath) != 0)\n") 
	fileWrite.write(threeTab + "return false;\n\n") 

	for index , item in enumerate(types):
		item_type = GetType(item)
		if item_type == g_structType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			fileWrite.write(twoTab + "size_t index_" + structName + " = csv.GetIndex(\"" + datas[index] + "\", " + str(g_rowName) + ");\n") 
			fileWrite.write(twoTab + "MsgAssert_Re0(index_" + structName + " != (size_t)-1 , \"error " + datas[index] + "\");\n\n") 
		elif item_type == g_structArrayType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			fileWrite.write(twoTab + "size_t index_" + structName + " = csv.GetIndex(\"" + datas[index] + "\", " + str(g_rowName) + ");\n") 
			fileWrite.write(twoTab + "MsgAssert_Re0(index_" + structName + " != (size_t)-1 , \"error " + datas[index] + "\");\n\n") 
		else:
			fileWrite.write(twoTab + "size_t index_" + datas[index] + " = csv.GetIndex(\"" + datas[index] + "\", " + str(g_rowName) + ");\n") 
			fileWrite.write(twoTab + "MsgAssert_Re0(index_" + datas[index] + " != (size_t)-1 , \"error " + datas[index] + "\");\n\n") 

	fileWrite.write(twoTab + "for (size_t row = " + str(g_rowDataStart) + "; row < csv.Count(); ++row)\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + g_configPrefix + filename + " conf;\n\n") 
	for index , item in enumerate(types):
		Str = GetTypeFunc(item)
		if type(Str) != str :
			if Str == g_int32ArrayFunc:
				fileWrite.write(threeTab + "{\n") 
				fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
				fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
				fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
				fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
				fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back((INT32)CUtil::atoi(vals[i].c_str()));\n") 
				fileWrite.write(threeTab + "}\n\n") 
			elif Str == g_int64ArrayFunc:
				fileWrite.write(threeTab + "{\n") 
				fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
				fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
				fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
				fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
				fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back((INT64)CUtil::atoi(vals[i].c_str()));\n") 
				fileWrite.write(threeTab + "}\n\n") 
			elif Str == g_boolArrayFunc:
				fileWrite.write(threeTab + "{\n") 
				fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
				fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
				fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
				fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
				fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back((bool)CUtil::atoi(vals[i].c_str()));\n") 
				fileWrite.write(threeTab + "}\n\n") 
			elif Str == g_dateArrayFunc:
				fileWrite.write(threeTab + "{\n") 
				fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
				fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
				fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
				fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
				fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back(Timer::Date(vals[i]));\n") 
				#fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back(Timer::Date(vals[i], " + GetDateType(datas[index]) + "));\n") 
				fileWrite.write(threeTab + "}\n\n") 
			elif Str == g_doubleArrayFunc:
				fileWrite.write(threeTab + "{\n") 
				fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
				fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
				fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
				fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n") 
				fileWrite.write(fiveTab + "conf." + datas[index] + ".push_back((float)CUtil::atof(vals[i].c_str()));\n") 
				fileWrite.write(threeTab + "}\n\n") 
			elif Str == g_stringArrayFunc:
				fileWrite.write(threeTab + "{\n") 
				fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + datas[index] + ");\n") 
				fileWrite.write(fourTab + "CUtil::tokenize(__tmp, conf." + datas[index] + ", \",\", \"\", \"\\\"\");\n") 
				fileWrite.write(threeTab + "}\n\n") 
			elif Str == g_structFunc:
				fileWrite.write(threeTab + "{\n") 

				npos = datas[index].find('[')
				structName = datas[index][0 : npos]
				structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
				childItems = item.split(',')

				fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
				fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + structName + ");\n") 
				fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \",\", \"\", \"\\\"\");\n") 
				fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n")
				fileWrite.write(fourTab + "{\n") 
				for indexChild , childItem in enumerate(childItems):
					fileWrite.write(fiveTab + "if(i == " + str(indexChild) + ")\n")
					fileWrite.write(fiveTab + "{\n") 

					childItemType = GetType(childItem)
					if childItemType == g_boolType:
						strVal = "bool val = "
						strVal = strVal + "CUtil::strtobool(vals[i].c_str()) >= 1;"
					elif childItemType == g_int32Type:
						strVal = "INT32 val = "
						strVal = strVal + "(INT32)CUtil::atoi(vals[i].c_str());"
					elif childItemType == g_dateType:
						strVal = "Timer::Date val(vals[i]);"
					elif childItemType == g_int64Type:
						strVal = "INT64 val = "
						strVal = strVal + "(INT64)CUtil::atoi(vals[i].c_str());"
					elif childItemType == g_doubleType:
						strVal = "double val = "
						strVal = strVal + "(float)CUtil::atof(vals[i].c_str());"
					elif childItemType == g_stringType:
						strVal = "std::string val = "
						strVal = strVal + "vals[i].c_str();"
					
					fileWrite.write(sixTab + strVal + "\n") 
					fileWrite.write(sixTab + "conf." + structName + "." + structDatas[indexChild] + " = val;\n") 

					fileWrite.write(fiveTab + "}\n") 
				fileWrite.write(fourTab + "}\n") 
				fileWrite.write(threeTab + "}\n\n") 

			elif Str == g_structArrayFunc:		
				fileWrite.write(threeTab + "{\n") 

				npos = datas[index].find('[')
				structName = datas[index][0 : npos]
				structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
				childItems = item.replace('[' , '').replace(']' , '').split(',')
				
				fileWrite.write(fourTab + "std::vector<std::string> vals;\n") 
				fileWrite.write(fourTab + "std::string __tmp = csv.GetString(row, index_" + structName + ");\n") 
				fileWrite.write(fourTab + "CUtil::tokenize(__tmp, vals, \"]\", \"\", \"\\\"\");\n") 
				fileWrite.write(fourTab + "for (size_t i = 0; i < vals.size(); ++i)\n")
				fileWrite.write(fourTab + "{\n") 
				fileWrite.write(fiveTab + "std::string strVal = vals[i];\n") 
				fileWrite.write(fiveTab + "if (strVal[0] == '[')\n") 
				fileWrite.write(sixTab + "strVal.assign(vals[i], 1, vals[i].length() - 1);\n\n") 
				fileWrite.write(fiveTab + g_configPrefix + filename + "::" + g_configPrefix + structName + "	" + "array;\n") 
				fileWrite.write(fiveTab + "std::vector<std::string> vals2;\n") 
				fileWrite.write(fiveTab + "CUtil::tokenize(strVal, vals2, \",\", \"\", \"\\\"\");\n") 
				fileWrite.write(fiveTab + "for (size_t j = 0; j < vals2.size(); ++j)\n") 
				fileWrite.write(fiveTab + "{\n")
				#childItems = re.split('(\[.*\])' , item)
				for indexChild , childItem in enumerate(childItems):
					fileWrite.write(sixTab + "if(j == " + str(indexChild) + ")\n")
					fileWrite.write(sixTab + "{\n") 

					childItemType = GetType(childItem)
					if childItemType == g_boolType:
						strVal = "bool val = "
						strVal = strVal + "CUtil::strtobool(vals2[j].c_str()) >= 1;"
					elif childItemType == g_int32Type:
						strVal = "INT32 val = "
						strVal = strVal + "(INT32)CUtil::atoi(vals2[j].c_str());"
					elif childItemType == g_dateType:
						strVal = "Timer::Date val(vals2[j]);"
					elif childItemType == g_int64Type:
						strVal = "INT64 val = "
						strVal = strVal + "(INT64)CUtil::atoi(vals2[j].c_str());"
					elif childItemType == g_doubleType:
						strVal = "double val = "
						strVal = strVal + "(float)CUtil::atof(vals2[j].c_str());"
					elif childItemType == g_stringType:
						strVal = "std::string val = "
						strVal = strVal + "vals2[j].c_str();"
					
					fileWrite.write(sixTab + oneTab + strVal + "\n") 
					fileWrite.write(sixTab + oneTab  + "array." + structDatas[indexChild] + " = val;\n") 

					fileWrite.write(sixTab + "}\n") 
				fileWrite.write(fiveTab + "}\n") 
				fileWrite.write(fiveTab + "conf.vec" + structName + ".push_back(array);\n") 
				fileWrite.write(fourTab + "}\n") 
				fileWrite.write(threeTab + "}\n\n")
		else:
			fileWrite.write(threeTab + "conf." + datas[index] + " = csv." + Str + "(row , index_" + datas[index] + ");\n") 
			

	fileWrite.write(threeTab + "m_vtConfigs.push_back(conf);\n") 
	fileWrite.write(twoTab + "}\n\n") 
	fileWrite.write(twoTab + "return true;\n") 
	fileWrite.write(oneTab + "}\n\n") 
			
	fileWrite.write(oneTab + g_configPrefix + filename + " & " + filename + "::Get(size_t row)\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "return m_vtConfigs.at(row);\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write("}\n\n") 
	
	fileWrite.close()	

def GenerateConfigHeader(filename , types , datas , comments):
	outputPath = g_xlsExportCPPPath  + os.sep + filename + ".h"
	if CheckNeedDelete(outputPath , types , datas ):
		return
	#if os.path.exists(outputPath): 
	#	os.remove(outputPath)

	fileWrite = open(outputPath , "a")
	
	loadConfig = filename + g_loadConfigSuffix
	dataConfig = g_configPrefix + filename

	strTitle = MakeTitle(types , datas)
	fileWrite.write(strTitle)	
	WriteFileDescription(fileWrite , filename + ".h" , "csv读取文件")
	fileWrite.write("#ifndef __" + g_xlsNamespace + "_" + filename + "_define_h__\n")
	fileWrite.write("#define __" + g_xlsNamespace + "_" + filename +  "_define_h__\n") 
	fileWrite.write("#include \"" + loadConfig + ".h\"\n\n") 
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n\n") 

	fileWrite.write(oneTab + "struct " + dataConfig + "\n") 
	fileWrite.write(oneTab + "{\n") 	 
	GenerateStructData(fileWrite , types , datas , comments)
	fileWrite.write(oneTab + "};\n\n\n") 
			
	fileWrite.write(oneTab + "class " + filename + "\n") 
	fileWrite.write(oneTab + "{\n") 
	
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "typedef std_unordered_map<" + GetType(types[0]) + " , " + dataConfig + "> MapConfigsT;\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "bool LoadFrom(const std::string& filepath);\n")
	fileWrite.write(twoTab + dataConfig + " * Get" + filename + "(INT32 nIndex);\n\n")
	fileWrite.write(oneTab + "private:\n")
	fileWrite.write(twoTab + "MapConfigsT m_mapConfigs;\n\n")
	
	fileWrite.write(oneTab + "};\n") 			
	fileWrite.write(oneTab + "extern " + filename + " * g_p" + filename + ";\n") 


	fileWrite.write("}\n\n" + "#endif// end" + "  __" + g_xlsNamespace + "_" + filename + "_define_h__\n") 
	
	fileWrite.close()	 	

def GenerateConfigCpp(filename , types , datas , comments):
	outputPath = g_xlsExportCPPPath  + os.sep + filename + ".cpp"
	if CheckNeedDelete(outputPath , types , datas ):
		return
	#if os.path.exists(outputPath): 
	#	os.remove(outputPath)

	loadConfig = filename + g_loadConfigSuffix
	dataConfig = g_configPrefix + filename
	fileWrite = open(outputPath , "a")
	
	strTitle = MakeTitle(types , datas)
	fileWrite.write(strTitle)	
	WriteFileDescription(fileWrite , filename + ".cpp" , "csv读取数据文件实现")
	fileWrite.write("#include \"" + filename + ".h\"\n") 
	fileWrite.write("#include \"LogLib/inc/Log.h\"\n\n") 
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n") 
	fileWrite.write(oneTab + "bool " + filename + "::LoadFrom(const std::string & filepath)\n") 
	fileWrite.write(oneTab + "{\n") 
	
	fileWrite.write(twoTab + g_xlsNamespace + "::" + loadConfig + " loadConfig;\n") 
	fileWrite.write(twoTab + "MsgAssert_Re0(loadConfig.LoadFrom(filepath) , \"Error " + filename + "LoadFrom \" << filepath);\n\n") 
	
	fileWrite.write(twoTab + "for(size_t i = 0; i < loadConfig.Count(); ++i)\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + g_xlsNamespace + "::" + g_configPrefix + loadConfig + "& config = loadConfig.Get(i);\n") 
	fileWrite.write(threeTab + g_xlsNamespace + "::" + dataConfig + " data = {0};\n") 
	
	for index , item in enumerate(types):
		itemType = GetType(item)
		if itemType == g_structType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			childItems = item.split(',')

			fileWrite.write(threeTab + "{\n") 
			for indexChild , childItem in enumerate(childItems):
				fileWrite.write(fourTab + "data." + structName + "." + structDatas[indexChild] + " = config." + structName + "." + structDatas[indexChild] + ";\n") 
			fileWrite.write(threeTab + "}\n")

		elif itemType == g_structArrayType:			
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			childItems = item.split(',')

			fileWrite.write(threeTab + "{\n") 
			fileWrite.write(fourTab + "std::vector<" + g_configPrefix + loadConfig + "::" + g_configPrefix + structName + ">::iterator iter = config.vec" + structName + ".begin();\n") 
			fileWrite.write(fourTab + "std::vector<" + g_configPrefix + loadConfig + "::" + g_configPrefix + structName + ">::iterator end = config.vec" + structName + ".end();\n") 
			fileWrite.write(fourTab + "for (; iter != end;++iter)\n") 
			fileWrite.write(fourTab + "{\n") 
			fileWrite.write(fiveTab + g_configPrefix + filename + "::" + g_configPrefix + structName + " array;\n")
			for indexChild , childItem in enumerate(childItems):
				fileWrite.write(fiveTab + "array." + structDatas[indexChild] + " = iter->" + structDatas[indexChild] + ";\n")
			fileWrite.write(fiveTab + "data.vec" + structName + ".push_back(array);\n")
			fileWrite.write(fourTab + "}\n")
			fileWrite.write(threeTab + "}\n")
		else:
			fileWrite.write(threeTab + "data." + datas[index] + " = config." + datas[index] + ";\n") 

	fileWrite.write(threeTab + "m_mapConfigs.insert(std::make_pair(data." + datas[0] + ",data));\n") 
	fileWrite.write(twoTab + "}\n") 
	fileWrite.write(twoTab + "return true;\n") 
	fileWrite.write(oneTab + "}\n\n") 
			
	fileWrite.write(oneTab + dataConfig + " * " + filename + "::Get" + filename + "(INT32 nIndex)\n")
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "MapConfigsT::iterator iter = m_mapConfigs.find(nIndex);\n") 
	fileWrite.write(twoTab + "if(iter == m_mapConfigs.end())\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + "gWarniStream( \"" + filename + "::Get" + filename + " NotFound \" << nIndex);\n") 
	fileWrite.write(threeTab + "return NULL;\n") 
	fileWrite.write(twoTab + "}\n\n") 
	fileWrite.write(twoTab + "return &iter->second;\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write(oneTab + filename + " * g_p" + filename + " = NULL;\n") 
	fileWrite.write("}\n\n") 
	
	fileWrite.close()

def GenerateConfigManagerHeader():
	outputPath = g_xlsExportCPPPath  + os.sep + "ConfigManager.h"
	if os.path.exists(outputPath): 
		os.remove(outputPath) 

	fileWrite = open(outputPath , "a") 

	WriteFileDescription(fileWrite , "ConfigManager.h" , "ConfigManager文件声明")
	fileWrite.write("#ifndef __" + g_xlsNamespace + "_ConfigManager" + "_define_h__\n")
	fileWrite.write("#define __" + g_xlsNamespace + "_ConfigManager" +  "_define_h__\n") 
	fileWrite.write("#include \"CUtil/inc/Common.h\"\n\n") 
	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n\n") 
			
	fileWrite.write(oneTab + "class ConfigManager\n") 
	fileWrite.write(oneTab + "{\n") 
	
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "ConfigManager();\n")
	fileWrite.write(twoTab + "~ConfigManager();\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "static ConfigManager & GetInstance();\n\n")
	fileWrite.write(oneTab + "public:\n")
	fileWrite.write(twoTab + "INT32		Init(std::string  strCsvPath);\n")
	fileWrite.write(twoTab + "INT32		Cleanup();\n")
	fileWrite.write(oneTab + "private:\n")
	fileWrite.write(twoTab + "\n\n")
	fileWrite.write(oneTab + "};\n") 		


	fileWrite.write("}\n\n" + "#endif// end" + "  __" + g_xlsNamespace + "_ConfigManager_define_h__\n") 
	
	fileWrite.close()	 	


def GenerateConfigManagerCPP():
	outputPath = g_xlsExportCPPPath  + os.sep + "ConfigManager.cpp"
	if os.path.exists(outputPath): 
		os.remove(outputPath)

	fileWrite = open(outputPath , "a")
	
	WriteFileDescription(fileWrite , "ConfigManager.cpp" , "ConfigManager数据管理文件实现")
	fileWrite.write("#include \"" + "ConfigManager.h\"\n") 
	fileWrite.write("#include \"LogLib/inc/Log.h\"\n") 
	for sheet , item in g_xlsRecords.items():
		fileWrite.write("#include \"" + sheet + ".h\"\n") 
	fileWrite.write(oneTab + "\n\n") 

	fileWrite.write("namespace " + g_xlsNamespace + "\n") 
	fileWrite.write("{\n") 
	fileWrite.write(oneTab + "ConfigManager::ConfigManager()\n") 
	fileWrite.write(oneTab + "{\n") 
	for sheet , item in g_xlsRecords.items():
		if sheet in g_xlsDeleteRecord:
			fileWrite.write("// " + twoTab + "g_p" + sheet + " = new " + g_xlsNamespace + "::" + sheet + ";\n") 
		else:
			fileWrite.write(twoTab + "g_p" + sheet + " = new " + g_xlsNamespace + "::" + sheet + ";\n") 
	fileWrite.write(oneTab + "}\n\n") 
	
	fileWrite.write(oneTab + "ConfigManager::~ConfigManager()\n") 
	fileWrite.write(oneTab + "{\n") 
	for sheet , item in g_xlsRecords.items():
		if sheet in g_xlsDeleteRecord:
			fileWrite.write("// " + twoTab + "SAFE_DELETE(" + g_xlsNamespace + "::" +  "g_p" + sheet + ");\n") 
		else:
			fileWrite.write(twoTab + "SAFE_DELETE(" + g_xlsNamespace + "::" +  "g_p" + sheet + ");\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write(oneTab + "ConfigManager & ConfigManager::GetInstance()\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "static ConfigManager instance;\n") 
	fileWrite.write(twoTab + "return instance;\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write(oneTab + "INT32 ConfigManager::Init(std::string  strCsvPath)\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "MsgAssert_ReF1(strCsvPath.length(), \"ConfigManager::Init error.\");\n\n") 
	fileWrite.write(twoTab + "if (strCsvPath[strCsvPath.length() - 1] != '/')\n") 
	fileWrite.write(twoTab + "{\n") 
	fileWrite.write(threeTab + "strCsvPath = strCsvPath + \"/\";\n") 
	fileWrite.write(twoTab + "}\n\n") 
	for sheet , item in g_xlsRecords.items():
		if sheet in g_xlsDeleteRecord:
			fileWrite.write("// " + twoTab + "MsgAssert_ReF1(" + g_xlsNamespace + "::" +  "g_p" + sheet + " , \"ConfigManager not Init\")\n") 
			fileWrite.write("// " + twoTab + "" + g_xlsNamespace + "::" +  "g_p" + sheet + "->LoadFrom(strCsvPath + \"" + sheet + ".tabcsv\");\n\n") 
		else:
			fileWrite.write(twoTab + "MsgAssert_ReF1(" + g_xlsNamespace + "::" +  "g_p" + sheet + " , \"ConfigManager not Init\")\n") 
			fileWrite.write(twoTab + "" + g_xlsNamespace + "::" +  "g_p" + sheet + "->LoadFrom(strCsvPath + \"" + sheet + ".tabcsv\");\n\n") 
	fileWrite.write(twoTab + "return 0;\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write(oneTab + "INT32 ConfigManager::Cleanup()\n") 
	fileWrite.write(oneTab + "{\n") 
	fileWrite.write(twoTab + "return -1;\n") 
	fileWrite.write(oneTab + "}\n\n") 

	fileWrite.write("}\n\n") 
	
	fileWrite.close()

################################流程相关函数处理#####################################
def RemoveNewLine(str):
	pp = ""
	for tt in str.splitlines():
		tt = tt.strip().lstrip().rstrip()
		pp = pp+tt
	return pp

def RemovListNewLine(List): 
	for index , item2 in enumerate(List):
		List[index] = RemoveNewLine(item2)

def GetType(item):
	if  item.lower() == "int".lower() or\
		item.lower() == "int32".lower():
		return g_int32Type
	elif item.lower() == "int64".lower() or\
		item.lower() == "long".lower():
		return g_int64Type
	elif item.lower() == "double".lower() or\
		item.lower() == "float".lower():
		return g_doubleType
	elif item.lower() == "bool".lower():
		return g_boolType
	elif item.lower() == "string".lower():
		return g_stringType
	elif item.lower() == "condition".lower():
		return g_conditionType
	elif item.lower() == "date".lower():
		return g_dateType
	elif item.lower() == "int[]".lower() or\
		item.lower() == "[]int".lower() or\
		item.lower() == "[]int32".lower() or\
		item.lower() == "int32[]".lower():
		return g_int32ArrayType
	elif item.lower() == "int64[]".lower() or\
		item.lower() == "[]int64".lower() or\
		item.lower() == "[]long".lower() or\
		item.lower() == "long[]".lower():
		return g_int64ArrayType
	elif item.lower() == "double[]".lower() or\
		item.lower() == "[]double".lower() or\
		item.lower() == "[]float".lower() or\
		item.lower() == "float[]".lower():
		return g_doubleArrayType
	elif item.lower() == "date[]".lower() or\
		item.lower() == "[]date".lower():
		return g_dateArrayType
	elif item.lower() == "string[]".lower() or\
		item.lower() == "[]string".lower() or\
		item.lower() == "[]std::string".lower() or\
		item.lower() == "std::string[]".lower():
		return g_stringArrayType
	elif item.lower() == "bool[]".lower() or\
		item.lower() == "[]bool".lower():
		return g_boolArrayType
	elif item.lower() == "Condition".lower():
		return g_boolArrayType
	elif item.lower().find(',') >= 0 and \
		item.lower()[0] != "[" and\
		item.lower()[len(item.lower()) - 1] != "]":
		return g_structType
	elif item.lower().find(',') >= 0 and \
		item.lower()[0] == "[" and\
		item.lower()[len(item.lower()) - 1] == "]":
		return g_structArrayType
	else:
		LogOutError("GetType error." , item)

	return "None"

def GetTypeFunc(item):
	if  item.lower() == "int".lower() or\
		item.lower() == "int32".lower():
		return g_int32Func
	elif item.lower() == "int64".lower() or\
		item.lower() == "long".lower():
		return g_int64Func
	elif item.lower() == "double".lower() or\
		item.lower() == "float".lower():
		return g_doubleFunc
	elif item.lower() == "bool".lower():
		return g_boolFunc
	elif item.lower() == "string".lower():
		return g_stringFunc
	elif item.lower() == "date".lower():
		return g_dateFunc
	elif item.lower() == "int[]".lower() or\
		item.lower() == "[]int".lower() or\
		item.lower() == "[]int32".lower() or\
		item.lower() == "int32[]".lower():
		return g_int32ArrayFunc
	elif item.lower() == "int64[]".lower() or\
		item.lower() == "[]int64".lower() or\
		item.lower() == "[]long".lower() or\
		item.lower() == "long[]".lower():
		return g_int64ArrayFunc
	elif item.lower() == "double[]".lower() or\
		item.lower() == "[]double".lower() or\
		item.lower() == "[]float".lower() or\
		item.lower() == "float[]".lower():
		return g_doubleArrayFunc
	elif item.lower() == "string[]".lower() or\
		item.lower() == "[]string".lower() or\
		item.lower() == "[]std::string".lower() or\
		item.lower() == "std::string[]".lower():
		return g_stringArrayFunc
	elif item.lower() == "bool[]".lower() or\
		item.lower() == "[]bool".lower():
		return g_boolArrayFunc
	elif item.lower() == "Date[]".lower() or\
		item.lower() == "[]Date".lower():
		return g_dateArrayFunc
	elif item.lower().index(",") >= 0 and \
		item.lower()[0] != "[" and\
		item.lower()[len(item.lower()) - 1] != "]":
		return g_structFunc
	elif item.lower().index(",") >= 0 and \
		item.lower()[0] == "[" and\
		item.lower()[len(item.lower()) - 1] == "]":
		return g_structArrayFunc
	else:
		LogOutError("GetTypeFunc error." , item)

	return None

def GetTypeTab(item):
	if  item.lower() == "int".lower() or\
		item.lower() == "int32".lower():
		return sixTab
	elif item.lower() == "bool".lower():
		return sixTab
	elif item.lower() == "int64".lower() or\
		item.lower() == "long".lower():
		return sixTab
	elif item.lower() == "double".lower() or\
		item.lower() == "float".lower():
		return sixTab
	elif item.lower() == "Date".lower():
		return fiveTab
	elif item.lower() == "string".lower():
		return fiveTab
	elif item.lower() == "int[]".lower() or\
		item.lower() == "[]int".lower() or\
		item.lower() == "[]int32".lower() or\
		item.lower() == "int32[]".lower():
		return threeTab
	elif item.lower() == "bool[]".lower() or\
		item.lower() == "[]bool".lower():
		return threeTab
	elif item.lower() == "int64[]".lower() or\
		item.lower() == "[]int64".lower() or\
		item.lower() == "[]long".lower() or\
		item.lower() == "long[]".lower():
		return threeTab
	elif item.lower() == "double[]".lower() or\
		item.lower() == "[]double".lower() or\
		item.lower() == "[]float".lower() or\
		item.lower() == "float[]".lower():
		return threeTab
	elif item.lower() == "string[]".lower() or\
		item.lower() == "[]string".lower() or\
		item.lower() == "[]std::string".lower() or\
		item.lower() == "std::string[]".lower():
		return oneTab
	elif item.lower() == "Date[]".lower() or\
		item.lower() == "[]Date".lower():
		return twoTab
	elif item.lower().index(",") >= 0 and \
		item.lower()[0] != "[" and\
		item.lower()[len(item.lower()) - 1] != "]":
		return twoTab
	elif item.lower().index(",") >= 0 and \
		item.lower()[0] == "[" and\
		item.lower()[len(item.lower()) - 1] == "]":
		return oneTab
	else:
		return twoTab

	return oneTab

def GetDateType(date):
	dateItems = date.split('-')
	if len(dateItems) == 4:
		return "Timer::DATE_TYPE_YEAR"
	elif len(dateItems) == 3:
		return "Timer::DATE_TYPE_MON"
	elif len(dateItems) == 2:
		return "Timer::DATE_TYPE_DAY"
	elif len(dateItems) == 1:
		timeItems = date.split(':')
		if len(timeItems) == 3:
			return "Timer::DATE_TYPE_HOUR"
		elif len(timeItems) == 2:
			return "Timer::DATE_TYPE_MIN"
		elif len(timeItems) == 1:
			return "Timer::DATE_TYPE_SEC"
	
	LogOutError("GetDateType error ." , date)
	return "Timer::DATE_TYPE_INVALID"
	
def RemoveSpecialWord(item):
	item = item.replace(' ','').replace('	','').replace('[','').replace(']','').strip().rstrip()
	if len(item) > 0 and item[len(item) - 1] == ',':
		#LogOutDebug("item:" , item)
		item = item[0:len(item) - 1]
		#LogOutDebug("end item:" , item)
	if len(item) > 0 and item[0] == ',':
		#LogOutDebug("item:" , item)
		item = item[1:]
		#LogOutDebug("end item:" , item)
	return item

def CheckDataArray(colItem):
	pp = ""
	childItems = colItem.split(',')
	for childIndex , childItem in enumerate(childItems):							
		childItem = RemoveSpecialWord(childItem)
		if len(childItem) > 0:
			pp = pp + childItem
			if len(childItems) - 1 != childIndex:
				pp = pp + ","

	#LogOutInfo("pp " , pp) 
	return pp

def CheckDataType(item_type , sheet , row , col , colItem):	
	if item_type == g_boolType:
		item = RemoveSpecialWord(colItem)
		if item.lower() == "true".lower() or\
			item.lower() == "false".lower() or\
			item.lower() == "0".lower() or\
			item.lower() == "1".lower():	
			pass
		else:
			LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , item  , " bool type error.")							

		if row != -1:
			g_xlsRecords[sheet][row][col] = item
		else:
			return item
	elif item_type == g_int32Type:
		item = RemoveSpecialWord(colItem)
		if not item.lower().isdigit():	
			LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , item  , " int32 type error.")							

		if row != -1:
			g_xlsRecords[sheet][row][col] = item
		else:
			return item
	elif item_type == g_int64Type:
		item = RemoveSpecialWord(colItem)
		if not item.lower().isdigit():	
			LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , item , " int64 type error.")							

		if row != -1:
			g_xlsRecords[sheet][row][col] = item
		else:
			return item
	elif item_type == g_doubleType:						
		item = RemoveSpecialWord(colItem)
		#if !item.lower().isdigit():	
		#	LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , " double type error.")							

		if row != -1:
			g_xlsRecords[sheet][row][col] = item
		else:
			return item
	elif item_type == g_stringType:
		return colItem
	elif item_type == g_conditionType:
		item = RemoveSpecialWord(colItem)
		if row != g_cellName and row != g_cellCS:
			tagitem = item.lower()[0:item.lower().find(':')]
			itemContent = item.lower()[item.lower().find(':') + 1:]
			#LogOutDebug("condition item:" , item , " tagitem=" , tagitem , " itemContent " , itemContent)
			if tagitem.lower() == g_conditionTag.lower() or tagitem.lower() == g_actionTag.lower():	
				pass
			else:
				LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , item , " condition tag error.")							

			if row != -1:
				g_xlsConditionRecords[sheet][col][row] = item
			else:
				return item
		return item

	elif item_type == g_dateType:
		item = RemoveSpecialWord(colItem)
		return item
	elif item_type == g_dateArrayType:
		g_xlsRecords[sheet][row][col] = CheckDataArray(colItem)
	elif item_type == g_int32ArrayType:
		g_xlsRecords[sheet][row][col] = CheckDataArray(colItem)
	elif item_type == g_int64ArrayType:
		g_xlsRecords[sheet][row][col] = CheckDataArray(colItem)
	elif item_type == g_doubleArrayType:
		g_xlsRecords[sheet][row][col] = CheckDataArray(colItem)
	elif item_type == g_dateArrayType:
		g_xlsRecords[sheet][row][col] = CheckDataArray(colItem)
	elif item_type == g_stringArrayType:
		pass
	elif item_type == g_conditionType:
		item = RemoveSpecialWord(colItem)
		if item.index("(") >= 0 or item.index(")") >= 0 or item.index("|") >= 0:	
			LogOutError("sheet=" , sheet , " :row=" , row , " :col=" , col , " item=" , item  , " Condition  type error.")							

		if row != -1:
			g_xlsRecords[sheet][row][col] = item
		else:
			return item
	elif item_type == g_structType:
#		if len(colItem) > 0 and colItem[len(colItem) - 1] == ',':
#			colItem = colItem[0:len(colItem) - 1]
#		if len(colItem) > 0 and colItem[0] == ',':
#			colItem = colItem[1:]
		childItems = colItem.split(',')		
		itemContent = ""
		for childIndex , childItem in enumerate(childItems):
			if len(childItem) == 0:
				continue
			childItem = RemoveSpecialWord(childItem)
			#LogOutDebug(" childItem" , childItem)
			childType = GetTypeByIndex(g_xlsRecords[sheet][g_rowType][col] , childIndex)
			#LogOutDebug("childType" , childType , " childItem" , childItem)
			childItem = CheckDataType(childType , sheet , -1 , col , childItem)
			itemContent = itemContent + childItem
			if len(childItems) - 1 != childIndex:
				itemContent = itemContent + ","

		#LogOutDebug("itemContent:" , itemContent)
		if row != -1:
			g_xlsRecords[sheet][row][col] = itemContent
		else:
			item = "["
			item = item + itemContent
			item = item + "]"
			return item

	elif item_type == g_structArrayType:
		if row != -1:
			childItems = colItem.split(']')		
			itemContent = ""
			for childIndex , childItem in enumerate(childItems):
				if len(childItem) > 0:
					#LogOutDebug("itemContent:" , itemContent , "childIndex" , childIndex , "size=" , len(childItems) , "childItems" , childItems , "colItem:" , colItem)
					childItem = childItem.replace('[' , '').replace(']' , '').strip().rstrip().lstrip()
					childItem = CheckDataType(g_structType , sheet , -1 , col , childItem)
					itemContent = itemContent + childItem

			g_xlsRecords[sheet][row][col] = itemContent
		else:
			return colItem
	else:
		return colItem

def GetTypeByIndex(types , index):
	typeItems = types.split(',')
	for childIndex , childItem in enumerate(typeItems):
		childItem = RemoveSpecialWord(childItem)
		if index == childIndex:
			return GetType(childItem)

	LogOutError("GetTypeByIndex error.types=" , types , " index=" ,index)

def MakeTitle(types , datas ):
	Str = "// attention dont't change this line:"		
	for index , item in enumerate(types):
		item_type = GetType(item)
		if item_type == g_structType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			childItems = item.split(',')
			Str = Str + structName
			if len(childItems) != len(structDatas):
				LogOutError("parase struct error.invalid size.index=" , index , " . name=" , structName , ":childItems " , childItems , "size=" , len(childItems) , ":structDatas " , structDatas , "size=" , len(structDatas))
			for indexChild , childItem in enumerate(childItems):
				Str = Str + GetType(childItem) + " " + structDatas[indexChild]  + ";"
		elif item_type == g_structArrayType:
			npos = datas[index].find('[')
			structName = datas[index][0 : npos]
			structDatas = datas[index][npos + 1 : len(datas[index]) - 1].split(',')
			
			Str = Str + structName
			item = item.replace('[' , '').replace(']' , '')
			childItems = item.split(',')
			if len(childItems) != len(structDatas):
				LogOutError("parase struct error.invalid size.index=" , index , " . name=" , structName , ":childItems " , childItems , "size=" , len(childItems) , ":structDatas " , structDatas , "size=" , len(structDatas))
			for indexChild , childItem in enumerate(childItems):
				Str = Str + GetType(childItem) + " " + structDatas[indexChild]  + ";"
		else:
			Str = Str + item_type + " " + datas[index]   + ";"
	Str += "\n"
	return Str
	
def CheckNeedDelete(outputFile , types , datas):
	if os.path.exists(outputFile): 
		strNew = MakeTitle(types , datas)
		
		fileRead = open(outputFile , "r")
		for line in fileRead: 
			strOld = line
			break

		#LogOutDebug("\nold comments:" , strOld , "new comments:" , strNew , "\n")
		if strNew != strOld:
			LogOutInfo(outputFile , " not match.\n")
			LogOutInfo("old comments:" , strOld)
			LogOutInfo("new comments:" , strNew)		
		fileRead.close()
		return True
	else:
		return False

################################流程无关函数处理#####################################
def Usage():
    print('GenerateCSV.py usage:')
    print('-h,--help: print help message.')
    print('-v, --version: print script version')
    print('-o, --output: input an csv output verb')
    print('-c, --output: input an cpp output verb')
    print('--foo: Test option ')
    print('--fre: another test option')

def Version():
	print('GenerateCSV.py 1.0.0.0.1')

def LogOutDebug(*string):
	cp = GetColor("debug")
	longStr = cp + "[ DEBUG ]"
	for item in range(len(string)):  
		longStr += str(string[item])

	print(longStr)
	GetColor("reset")

def LogOutInfo(*string):
	cp = GetColor("info")
	longStr = cp + "[ INFO ] "
	for item in range(len(string)):  
		longStr += str(string[item])
	
	print(longStr)
	GetColor("reset")
	
def LogOutError(*string):
	cp = GetColor("error")
	longStr = cp + "[ ERR ] "
	for item in range(len(string)):  
		longStr += str(string[item])
	
	print(longStr)
	GetColor("reset")
	sys.exit()
	
def InitColor():
	if g_platform == "LINUX":
		cp = '\033['
		g_yellow = cp + '33m'
		g_yellow_h = cp + '1;33m'
		g_green = cp + '32m'
		g_green_h = cp + '1;32m'
		g_red = cp + '31m'
		g_cyan = cp + '36m'
		g_original = cp + '0m'
	else:
		g_yellow = ""
		g_yellow_h = ""
		g_green = ""
		g_green_h = ""
		g_red = ""
		g_original = ""
		g_cyan = ""

def GetColor(type):
	stdOutHandle = ctypes.windll.kernel32.GetStdHandle(g_stdOutputHandle)
	if type == "error":
		if g_platform == "LINUX":
			return g_red
		else:
			ctypes.windll.kernel32.SetConsoleTextAttribute(stdOutHandle , FOREGROUND_RED | FOREGROUND_INTENSTITY)
			return ""
	elif type == "info":
		if g_platform == "LINUX":
			return g_green
		else:
			ctypes.windll.kernel32.SetConsoleTextAttribute(stdOutHandle , FOREGROUND_GREEN | FOREGROUND_INTENSTITY)
			return ""
	elif type == "debug" or type == "reset":
		if g_platform == "LINUX":
			return g_original
		else:
			ctypes.windll.kernel32.SetConsoleTextAttribute(stdOutHandle , FOREGROUND_BLUE | FOREGROUND_RED | FOREGROUND_GREEN)
			return ""
	elif type == "warning" :
		if g_platform == "LINUX":
			return g_yellow
		else:
			ctypes.windll.kernel32.SetConsoleTextAttribute(stdOutHandle , FOREGROUND_BLUE | FOREGROUND_INTENSTITY)
			return ""
				
def WriteFileDescription(fileWrite , sfile , desc):
	fileWrite.write("/************************************" + "\n")
	fileWrite.write("FileName	:	" + sfile + "\n")
	fileWrite.write("Author		:	generate by tools" + "\n")
	fileWrite.write("HostName	:	" + socket.gethostname() + "\n")
	fileWrite.write("IP			:	" + socket.gethostbyname(socket.gethostname()) + "\n")
	fileWrite.write("Version		:	0.0.1" + "\n")
	fileWrite.write("Date		:	" + time.strftime('%Y-%m-%d %H:%M:%S') + "\n")
	fileWrite.write("Description	:	" + desc + "\n")
	fileWrite.write("************************************/" + "\n")

def Search(base , suffix):
    pattern = suffix
    fileresult = []
    cur_list = os.listdir(base)
    for item in cur_list:
        full_path = os.path.join(base, item)
        if os.path.isdir(full_path):            
            fileresult += Search(full_path , suffix)
        if os.path:
            if full_path.endswith(pattern):
                fileresult.append(full_path)
    return fileresult
	
def DeleteExportPathFiles():
	files = Search(g_xlsExportCSVPath , '.tabcsv')
	for sfile in files:
		if os.path.exists(sfile): 
			os.remove(sfile)

def CreateExportPathFiles(): 
	if False == os.path.exists(g_xlsExportCSVPath):
		os.makedirs(g_xlsExportCSVPath)
		LogOutInfo("create dir: " , g_xlsExportCSVPath)
		
	if False == os.path.exists(g_xlsExportCPPPath):
		os.makedirs(g_xlsExportCPPPath)
		LogOutInfo("create dir: " , g_xlsExportCPPPath)


################################main函数处理#####################################
def handleArgs(argv): 
	global g_xlsImportPath 
	global g_xlsExportCSVPath
	global g_xlsExportCPPPath
	
	try:
		opts, args = getopt.getopt(argv[1:], 'hvi:o:c:', ['import='])
	except: 
		Usage()
		sys.exit(2) 
	if len(argv) < 2: 
		Usage()
		return  
	for o, a in opts:
		if o in ('-h', '--help'):
			Usage()
			sys.exit(1)
		elif o in ('-v', '--version'):
			Version()
			sys.exit(0) 
		elif o in ('-i','--import',):
			g_xlsImportPath = a
		elif o in ('-o','--export',):
			g_xlsExportCSVPath = a
		elif o in ('-c','--cppexport',):
			g_xlsExportCPPPath = a
		elif o in ('--fre',):
			Fre=a
		else:
			print('unhandled option')
			sys.exit(3) 
			
def main(argv):
	global g_xlsImportPath 
	global g_xlsExportCSVPath
	global g_xlsExportCPPPath
	InitColor()
	#handleArgs(argv)
	g_xlsImportPath = "./xls_config"
	g_xlsExportCSVPath = "../../../bin/vs14.0/x64/Lib_Debug_x64/csv_config"
	g_xlsExportCPPPath = "../../../vsproject/TestLibCore/CSVConfigs"
	LogOutInfo("generate csv from path:" + g_xlsImportPath + " csv will export to:" + g_xlsExportCSVPath + " cpp will export to:" + g_xlsExportCPPPath) 
	start()
	LogOutInfo("complete generate csv.") 
	
if __name__ == '__main__': 
	main(sys.argv)


